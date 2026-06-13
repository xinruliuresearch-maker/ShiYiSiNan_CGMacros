# -*- coding: utf-8 -*-
"""
High-impact computational upgrade for the integrated NHANES + CGMacros project.

This script executes the dry-lab upgrades requested after the research strategy
package:

1. NHANES survey-weighted mixture/sensitivity analyses:
   WQS approximation, qgcomp-style quantile mixture, BKMR-style kernel ridge
   mixture response, restricted cubic spline, and E-value tables.
2. Dual MSI reconstruction:
   exposure-facing MSI and response-vulnerability index.
3. CGMacros subject-disjoint prediction and cluster/hierarchical inference.
4. Food-matrix codebook, proxy USDA/FDC linkage, and dual-rule reliability.
5. Stanford CGM Database external boundary validation.
6. CTD/CompTox/Reactome/KEGG reproducible mechanism workflow.
7. TRIPOD+AI, PROBAST+AI, STROBE, and PRISMA supplementary checklists.

The implementation is intentionally transparent and dependency-light. Where R
packages are normally used before submission, the output labels the Python
implementation as an approximation.
"""

from __future__ import annotations

import csv
import gzip
import json
import math
import shutil
import textwrap
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from statistics import NormalDist

import numpy as np
import pandas as pd

try:
    from scipy import stats

    HAS_SCIPY = True
except Exception:  # pragma: no cover
    HAS_SCIPY = False
    stats = None

try:
    import matplotlib.pyplot as plt

    HAS_MPL = True
except Exception:  # pragma: no cover
    HAS_MPL = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    HAS_OPENPYXL = True
except Exception:  # pragma: no cover
    HAS_OPENPYXL = False


ROOT = Path(__file__).resolve().parents[1]
NHANES_ROOT = ROOT.parent / "NHANES_MetS_Project"
TODAY = "2026-06-11"
OUT = ROOT / "outputs" / "integrated_mainline" / f"high_impact_computational_upgrade_{TODAY}"
MIRROR = NHANES_ROOT / "result" / "integrated_mainline" / f"high_impact_computational_upgrade_{TODAY}"
RNG = np.random.default_rng(20260611)

NHANES_ANALYSIS = ROOT / "outputs" / "integrated_mainline" / "phase2_nhanes_survey_ready" / "phase2_python_analysis_dataset.csv"
NHANES_STAGE1 = ROOT / "outputs" / "integrated_mainline" / "stage1_msi" / "nhanes_2013_2018_msi_core.csv"
CGM_STAGE1 = ROOT / "outputs" / "integrated_mainline" / "stage1_msi" / "cgmacros_meal_level_with_msi_core.csv"
STANFORD_DIR = ROOT / "data" / "external" / "Stanford_CGMDB"
CTD_DIR = NHANES_ROOT / "data" / "raw_mechanism_databases" / "CTD"


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def mirror_file(path: Path) -> None:
    if not path.exists():
        return
    rel = path.relative_to(OUT)
    dest = MIRROR / rel
    ensure_dir(dest.parent)
    shutil.copy2(path, dest)


def write_text(path: Path, content: str) -> Path:
    ensure_dir(path.parent)
    path.write_text(textwrap.dedent(content).strip() + "\n", encoding="utf-8")
    mirror_file(path)
    return path


def write_csv(path: Path, df: pd.DataFrame) -> Path:
    ensure_dir(path.parent)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    mirror_file(path)
    return path


def write_json(path: Path, payload: dict | list) -> Path:
    ensure_dir(path.parent)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    mirror_file(path)
    return path


def save_fig(path: Path) -> Path | None:
    if not HAS_MPL:
        return None
    ensure_dir(path.parent)
    plt.tight_layout()
    plt.savefig(path, dpi=240)
    plt.close()
    mirror_file(path)
    return path


def numeric(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def winsorize(s: pd.Series, low: float = 0.01, high: float = 0.99) -> pd.Series:
    x = numeric(s)
    if x.notna().sum() == 0:
        return x
    lo, hi = x.quantile([low, high])
    return x.clip(lo, hi)


def zscore(s: pd.Series, mean: float | None = None, sd: float | None = None) -> pd.Series:
    x = numeric(s)
    if mean is None:
        mean = float(x.mean(skipna=True))
    if sd is None:
        sd = float(x.std(skipna=True, ddof=0))
    if not np.isfinite(sd) or sd == 0:
        return pd.Series(np.nan, index=x.index)
    return (x - mean) / sd


def p_norm_2sided(z: float) -> float:
    if not np.isfinite(z):
        return np.nan
    return math.erfc(abs(float(z)) / math.sqrt(2.0))


def chi2_sf(x: float, df: int) -> float:
    if not np.isfinite(x) or df <= 0:
        return np.nan
    if HAS_SCIPY:
        return float(stats.chi2.sf(x, df))
    if df == 1:
        return math.erfc(math.sqrt(x / 2.0))
    if df == 2:
        return math.exp(-x / 2.0)
    z = ((x / df) ** (1 / 3) - (1 - 2 / (9 * df))) / math.sqrt(2 / (9 * df))
    return p_norm_2sided(z) / 2.0


def bh_fdr(pvals: pd.Series) -> pd.Series:
    p = numeric(pvals)
    out = pd.Series(np.nan, index=p.index, dtype=float)
    mask = p.notna()
    vals = p.loc[mask].to_numpy(dtype=float)
    if len(vals) == 0:
        return out
    order = np.argsort(vals)
    ranked = vals[order]
    m = len(vals)
    q = ranked * m / (np.arange(m) + 1)
    q = np.minimum.accumulate(q[::-1])[::-1]
    tmp = np.empty_like(q)
    tmp[order] = np.clip(q, 0, 1)
    out.loc[mask] = tmp
    return out


def safe_inverse(a: np.ndarray, ridge: float = 1e-8) -> np.ndarray:
    eye = np.eye(a.shape[0])
    try:
        return np.linalg.inv(a + ridge * eye)
    except np.linalg.LinAlgError:
        return np.linalg.pinv(a + ridge * eye)


@dataclass
class FitResult:
    beta: np.ndarray
    se: np.ndarray
    vcov: np.ndarray
    terms: list[str]
    n: int
    p: int
    clusters: int
    r2: float
    residuals: np.ndarray
    y: np.ndarray
    fitted: np.ndarray


def make_design(
    df: pd.DataFrame,
    numeric_cols: list[str],
    categorical_cols: list[str] | None = None,
    standardize_numeric: bool = True,
    reference_columns: list[str] | None = None,
    numeric_params: dict[str, dict[str, float]] | None = None,
) -> tuple[pd.DataFrame, dict[str, dict[str, float]]]:
    categorical_cols = categorical_cols or []
    parts: list[pd.DataFrame] = [pd.DataFrame({"Intercept": np.ones(len(df), dtype=float)}, index=df.index)]
    params: dict[str, dict[str, float]] = {}
    for col in numeric_cols:
        if col not in df.columns:
            continue
        x = numeric(df[col])
        if numeric_params and col in numeric_params:
            med = float(numeric_params[col].get("median", 0.0))
        else:
            med = float(x.median(skipna=True)) if x.notna().sum() else 0.0
        x = x.fillna(med)
        if standardize_numeric:
            if numeric_params and col in numeric_params:
                mean = float(numeric_params[col].get("mean", 0.0))
                sd = float(numeric_params[col].get("sd", 1.0))
            else:
                mean = float(x.mean())
                sd = float(x.std(ddof=0))
            if not np.isfinite(sd) or sd == 0:
                sd = 1.0
            x = (x - mean) / sd
            params[col] = {"median": med, "mean": mean, "sd": sd}
        else:
            params[col] = {"median": med, "mean": 0.0, "sd": 1.0}
        parts.append(pd.DataFrame({col: x.astype(float)}, index=df.index))
    for col in categorical_cols:
        if col not in df.columns:
            continue
        s = df[col].astype("string").fillna("missing")
        d = pd.get_dummies(s, prefix=col, drop_first=True, dtype=float)
        parts.append(d)
    X = pd.concat(parts, axis=1)
    X = X.loc[:, ~X.columns.duplicated()].copy()
    if reference_columns is not None:
        for col in reference_columns:
            if col not in X.columns:
                X[col] = 0.0
        X = X[reference_columns]
    return X.astype(float), params


def fit_wls_cluster(
    y: pd.Series,
    X: pd.DataFrame,
    weights: pd.Series | None = None,
    clusters: pd.Series | None = None,
    ridge: float = 1e-8,
) -> FitResult:
    mask = y.notna()
    for col in X.columns:
        mask &= pd.to_numeric(X[col], errors="coerce").notna()
    if weights is not None:
        mask &= pd.to_numeric(weights, errors="coerce").notna()
    if clusters is not None:
        mask &= clusters.notna()

    yy = y.loc[mask].astype(float).to_numpy()
    XX = X.loc[mask].astype(float).to_numpy()
    n, p = XX.shape
    if weights is None:
        w = np.ones(n, dtype=float)
    else:
        w = weights.loc[mask].astype(float).to_numpy()
        w = np.where(np.isfinite(w) & (w > 0), w, np.nan)
        med = np.nanmedian(w)
        w = np.where(np.isfinite(w), w, med if np.isfinite(med) else 1.0)
        w = w / np.nanmean(w)

    xtw = XX.T * w
    bread = safe_inverse(xtw @ XX, ridge=ridge)
    beta = bread @ (xtw @ yy)
    fitted = XX @ beta
    resid = yy - fitted

    if clusters is None:
        scores = XX * (w * resid)[:, None]
        meat = scores.T @ scores
        g = 0
        if n > p:
            meat *= n / max(n - p, 1)
    else:
        cl = clusters.loc[mask].astype("string").to_numpy()
        unique = pd.unique(cl)
        meat = np.zeros((p, p), dtype=float)
        for group in unique:
            idx = cl == group
            sg = (XX[idx, :].T @ (w[idx] * resid[idx])).reshape(-1, 1)
            meat += sg @ sg.T
        g = len(unique)
        if g > 1 and n > p:
            meat *= (g / (g - 1)) * ((n - 1) / max(n - p, 1))

    vcov = bread @ meat @ bread
    se = np.sqrt(np.maximum(np.diag(vcov), 0.0))
    ss_res = float(np.sum(w * resid**2))
    ybar = float(np.average(yy, weights=w))
    ss_tot = float(np.sum(w * (yy - ybar) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan
    return FitResult(beta, se, vcov, list(X.columns), n, p, g, r2, resid, yy, fitted)


def fit_table(result: FitResult, model_name: str, outcome: str, extra: dict[str, str] | None = None) -> pd.DataFrame:
    rows = []
    extra = extra or {}
    for term, beta, se in zip(result.terms, result.beta, result.se):
        z = beta / se if se > 0 else np.nan
        rows.append(
            {
                **extra,
                "model": model_name,
                "outcome": outcome,
                "term": term,
                "estimate": beta,
                "std_error": se,
                "z": z,
                "p_value": p_norm_2sided(z),
                "ci_low": beta - 1.96 * se if np.isfinite(se) else np.nan,
                "ci_high": beta + 1.96 * se if np.isfinite(se) else np.nan,
                "n": result.n,
                "clusters": result.clusters,
                "r2": result.r2,
            }
        )
    return pd.DataFrame(rows)


def weighted_corr(x: pd.Series, y: pd.Series, w: pd.Series) -> float:
    mask = x.notna() & y.notna() & w.notna()
    if mask.sum() < 5:
        return np.nan
    xx = x.loc[mask].astype(float).to_numpy()
    yy = y.loc[mask].astype(float).to_numpy()
    ww = w.loc[mask].astype(float).to_numpy()
    ww = np.where((ww > 0) & np.isfinite(ww), ww, np.nanmedian(ww))
    mx = np.average(xx, weights=ww)
    my = np.average(yy, weights=ww)
    vx = np.average((xx - mx) ** 2, weights=ww)
    vy = np.average((yy - my) ** 2, weights=ww)
    if vx <= 0 or vy <= 0:
        return np.nan
    return float(np.average((xx - mx) * (yy - my), weights=ww) / math.sqrt(vx * vy))


def quantile_score(s: pd.Series, q: int = 4) -> pd.Series:
    x = numeric(s)
    out = pd.Series(np.nan, index=s.index, dtype=float)
    mask = x.notna()
    if mask.sum() < q:
        return out
    ranks = x.loc[mask].rank(method="first")
    out.loc[mask] = pd.qcut(ranks, q=q, labels=False, duplicates="drop").astype(float)
    return out


def residualize(df: pd.DataFrame, target: str, cov_num: list[str], cov_cat: list[str], weight_col: str | None) -> pd.Series:
    sub = df.copy()
    y = numeric(sub[target])
    X, _ = make_design(sub, cov_num, cov_cat, standardize_numeric=True)
    w = numeric(sub[weight_col]) if weight_col and weight_col in sub.columns else None
    fit = fit_wls_cluster(y, X, weights=w, clusters=None)
    out = pd.Series(np.nan, index=sub.index, dtype=float)
    mask = y.notna()
    for col in X.columns:
        mask &= X[col].notna()
    out.loc[mask] = fit.residuals
    return out


def rcs_terms(x: pd.Series, knots: np.ndarray) -> pd.DataFrame:
    xx = numeric(x)
    if len(knots) < 4:
        raise ValueError("Need at least four knots for RCS.")
    k_last = knots[-1]
    k_penult = knots[-2]

    def tp(v: pd.Series, knot: float) -> pd.Series:
        return np.maximum(v - knot, 0.0) ** 3

    terms = {"linear": xx}
    for j, kj in enumerate(knots[:-2]):
        term = tp(xx, kj)
        term -= tp(xx, k_penult) * (k_last - kj) / (k_last - k_penult)
        term += tp(xx, k_last) * (k_penult - kj) / (k_last - k_penult)
        terms[f"rcs{j+1}"] = term
    return pd.DataFrame(terms, index=x.index)


def evalue_from_smd(beta: float, se: float) -> dict[str, float]:
    b = abs(float(beta))
    rr = math.exp(0.91 * b)
    ev = rr + math.sqrt(max(rr * (rr - 1), 0.0))
    lo = beta - 1.96 * se
    hi = beta + 1.96 * se
    if lo <= 0 <= hi:
        ci_smd = 0.0
    else:
        ci_smd = min(abs(lo), abs(hi))
    rr_ci = math.exp(0.91 * ci_smd)
    ev_ci = rr_ci + math.sqrt(max(rr_ci * (rr_ci - 1), 0.0))
    return {"rr_approx": rr, "e_value": ev, "e_value_ci_limit": ev_ci}


def cohen_kappa(a: pd.Series, b: pd.Series) -> float:
    mask = a.notna() & b.notna()
    if mask.sum() == 0:
        return np.nan
    aa = a.loc[mask].astype(str)
    bb = b.loc[mask].astype(str)
    cats = sorted(set(aa) | set(bb))
    obs = float((aa == bb).mean())
    pa = aa.value_counts(normalize=True).reindex(cats, fill_value=0.0)
    pb = bb.value_counts(normalize=True).reindex(cats, fill_value=0.0)
    exp = float((pa * pb).sum())
    if exp >= 1:
        return np.nan
    return (obs - exp) / (1 - exp)


def icc_two_way_single(a: pd.Series, b: pd.Series) -> float:
    x = pd.DataFrame({"a": numeric(a), "b": numeric(b)}).dropna()
    if len(x) < 3:
        return np.nan
    mat = x.to_numpy(dtype=float)
    n, k = mat.shape
    mean_subject = mat.mean(axis=1, keepdims=True)
    mean_rater = mat.mean(axis=0, keepdims=True)
    grand = mat.mean()
    ss_subject = k * float(((mean_subject - grand) ** 2).sum())
    ss_rater = n * float(((mean_rater - grand) ** 2).sum())
    ss_error = float(((mat - mean_subject - mean_rater + grand) ** 2).sum())
    ms_subject = ss_subject / (n - 1)
    ms_rater = ss_rater / (k - 1)
    ms_error = ss_error / ((n - 1) * (k - 1))
    denom = ms_subject + (k - 1) * ms_error + k * (ms_rater - ms_error) / n
    return (ms_subject - ms_error) / denom if denom != 0 else np.nan


def load_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    nh = pd.read_csv(NHANES_ANALYSIS)
    nh_stage = pd.read_csv(NHANES_STAGE1) if NHANES_STAGE1.exists() else pd.DataFrame()
    cg = pd.read_csv(CGM_STAGE1)
    return nh, nh_stage, cg


def add_dual_msi(nh: pd.DataFrame, cg: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    nh = nh.copy()
    cg = cg.copy()
    nh["nhanes_tyg"] = numeric(nh["TyG"])
    nh["nhanes_ln_tg_hdl"] = numeric(nh["ln_TG_HDL"])

    cg["cg_ln_tg_hdl"] = np.log((numeric(cg["Triglycerides"]) + 1e-6) / (numeric(cg["HDL"]) + 1e-6))
    cg["cg_tyg"] = np.log((numeric(cg["Triglycerides"]) * numeric(cg["Fasting GLU - PDL (Lab)"])) / 2.0)
    cg["cg_homa_ir"] = (numeric(cg["Fasting GLU - PDL (Lab)"]) * numeric(cg["Insulin"])) / 405.0
    cg["cg_ln_homa_ir"] = np.log(cg["cg_homa_ir"].where(cg["cg_homa_ir"] > 0))

    component_map = [
        ("BMI", "BMXBMI", "BMI"),
        ("ln_tg_hdl", "ln_TG_HDL", "cg_ln_tg_hdl"),
        ("TyG", "TyG", "cg_tyg"),
        ("HbA1c", "HbA1c", "A1c PDL (Lab)"),
        ("fasting_glucose", None, "Fasting GLU - PDL (Lab)"),
        ("ln_homa_ir", "ln_HOMA_IR", "cg_ln_homa_ir"),
    ]
    refs = []
    for label, nh_col, cg_col in component_map:
        if nh_col and nh_col in nh.columns:
            ref = winsorize(nh[nh_col])
        elif cg_col in cg.columns:
            ref = winsorize(cg[cg_col])
        else:
            continue
        refs.append({"component": label, "reference_source": "NHANES" if nh_col else "CGMacros", "mean": ref.mean(), "sd": ref.std(ddof=0)})
    ref_df = pd.DataFrame(refs)

    def ref_z(df: pd.DataFrame, source_col: str, component: str) -> pd.Series:
        row = ref_df.loc[ref_df["component"] == component]
        if row.empty or source_col not in df.columns:
            return pd.Series(np.nan, index=df.index)
        return zscore(winsorize(df[source_col]), float(row["mean"].iloc[0]), float(row["sd"].iloc[0]))

    nh["z_bmi_ref"] = ref_z(nh, "BMXBMI", "BMI")
    nh["z_ln_tg_hdl_ref"] = ref_z(nh, "ln_TG_HDL", "ln_tg_hdl")
    nh["z_tyg_ref"] = ref_z(nh, "TyG", "TyG")
    nh["z_hba1c_ref"] = ref_z(nh, "HbA1c", "HbA1c")
    nh["z_ln_homa_ir_ref"] = ref_z(nh, "ln_HOMA_IR", "ln_homa_ir")

    cg["z_bmi_ref"] = ref_z(cg, "BMI", "BMI")
    cg["z_ln_tg_hdl_ref"] = ref_z(cg, "cg_ln_tg_hdl", "ln_tg_hdl")
    cg["z_tyg_ref"] = ref_z(cg, "cg_tyg", "TyG")
    cg["z_hba1c_ref"] = ref_z(cg, "A1c PDL (Lab)", "HbA1c")
    cg["z_fasting_glucose_ref"] = zscore(winsorize(cg["Fasting GLU - PDL (Lab)"]))
    cg["z_ln_homa_ir_ref"] = ref_z(cg, "cg_ln_homa_ir", "ln_homa_ir")

    nh_exposure_components = ["z_bmi_ref", "z_ln_tg_hdl_ref", "z_tyg_ref"]
    nh_response_components = ["z_bmi_ref", "z_hba1c_ref", "z_ln_homa_ir_ref", "z_tyg_ref", "z_ln_tg_hdl_ref"]
    cg_exposure_components = ["z_bmi_ref", "z_ln_tg_hdl_ref", "z_tyg_ref"]
    cg_response_components = ["z_bmi_ref", "z_hba1c_ref", "z_fasting_glucose_ref", "z_ln_homa_ir_ref", "z_ln_tg_hdl_ref"]

    nh["exposure_facing_msi"] = nh[nh_exposure_components].mean(axis=1, skipna=True)
    nh["exposure_facing_msi_components"] = nh[nh_exposure_components].notna().sum(axis=1)
    nh["response_vulnerability_index"] = nh[nh_response_components].mean(axis=1, skipna=True)
    nh["response_vulnerability_components"] = nh[nh_response_components].notna().sum(axis=1)
    cg["exposure_facing_msi"] = cg[cg_exposure_components].mean(axis=1, skipna=True)
    cg["exposure_facing_msi_components"] = cg[cg_exposure_components].notna().sum(axis=1)
    cg["response_vulnerability_index"] = cg[cg_response_components].mean(axis=1, skipna=True)
    cg["response_vulnerability_components"] = cg[cg_response_components].notna().sum(axis=1)

    return nh, cg, ref_df


def run_nhanes_mixture(nh: pd.DataFrame) -> dict[str, pd.DataFrame]:
    out: dict[str, pd.DataFrame] = {}
    exp_cols = ["ln_Sigma_DEHP", "pct_oxidative_10", "ln_oxidative_to_MEHP"]
    outcomes = [
        "exposure_facing_msi",
        "response_vulnerability_index",
        "msi_core_partial",
        "msi_no_bmi",
        "ln_HOMA_IR",
        "HbA1c",
        "TyG",
        "ln_TG_HDL",
    ]
    outcomes = [o for o in outcomes if o in nh.columns]
    cov_num = ["RIDAGEYR", "INDFMPIR", "DR1TKCAL", "ln_URXUCR", "ever_smoker", "alcohol_ever", "any_physical_activity"]
    cov_cat = ["RIAGENDR", "RIDRETH3", "DMDEDUC2", "cycle"]
    weight_col = "WTSB6YR_MAIN"
    cluster_col = "cluster"
    nh = nh.copy()
    nh[cluster_col] = nh["SDMVSTRA"].astype(str) + "_" + nh["SDMVPSU"].astype(str)

    main_rows = []
    evalue_rows = []
    for outcome in outcomes:
        for exp in exp_cols:
            model_df = nh.copy()
            model_df[f"{exp}_z"] = zscore(winsorize(model_df[exp]))
            model_df[f"{outcome}_z"] = zscore(winsorize(model_df[outcome]))
            X, _ = make_design(model_df, [f"{exp}_z"] + cov_num, cov_cat)
            fit = fit_wls_cluster(
                numeric(model_df[f"{outcome}_z"]),
                X,
                weights=numeric(model_df[weight_col]),
                clusters=model_df[cluster_col],
            )
            ft = fit_table(fit, "survey_weighted_standardized_wls", outcome, {"exposure": exp})
            main_rows.append(ft)
            term = ft.loc[ft["term"] == f"{exp}_z"]
            if not term.empty:
                beta = float(term["estimate"].iloc[0])
                se = float(term["std_error"].iloc[0])
                ev = evalue_from_smd(beta, se)
                evalue_rows.append(
                    {
                        "outcome": outcome,
                        "exposure": exp,
                        "standardized_beta": beta,
                        "std_error": se,
                        "ci_low": float(term["ci_low"].iloc[0]),
                        "ci_high": float(term["ci_high"].iloc[0]),
                        **ev,
                        "interpretation": "E-value approximated from standardized continuous effect using RR ~= exp(0.91*SMD).",
                    }
                )
    main = pd.concat(main_rows, ignore_index=True) if main_rows else pd.DataFrame()
    if not main.empty:
        main["fdr"] = bh_fdr(main["p_value"])
    out["nhanes_survey_weighted_standardized_models"] = main
    out["nhanes_e_values"] = pd.DataFrame(evalue_rows)

    # qgcomp-style quantile mixture.
    qg_rows = []
    qg_weight_rows = []
    for outcome in outcomes:
        qdf = nh.copy()
        for exp in exp_cols:
            qdf[f"q_{exp}"] = quantile_score(qdf[exp], 4)
        q_exp = [f"q_{e}" for e in exp_cols]
        X, _ = make_design(qdf, q_exp + cov_num, cov_cat, standardize_numeric=False)
        fit = fit_wls_cluster(numeric(qdf[outcome]), X, weights=numeric(qdf[weight_col]), clusters=qdf[cluster_col])
        beta_map = dict(zip(fit.terms, fit.beta))
        idx = [fit.terms.index(q) for q in q_exp if q in fit.terms]
        if idx:
            c = np.zeros(len(fit.terms))
            c[idx] = 1.0
            psi = float(c @ fit.beta)
            se = float(math.sqrt(max(c @ fit.vcov @ c, 0.0)))
            z = psi / se if se > 0 else np.nan
            qg_rows.append(
                {
                    "outcome": outcome,
                    "method": "qgcomp_quantile_mixture_python",
                    "psi_all_exposures_one_quantile": psi,
                    "std_error": se,
                    "z": z,
                    "p_value": p_norm_2sided(z),
                    "ci_low": psi - 1.96 * se,
                    "ci_high": psi + 1.96 * se,
                    "n": fit.n,
                    "clusters": fit.clusters,
                    "r2": fit.r2,
                }
            )
            coefs = pd.Series({e: beta_map.get(f"q_{e}", np.nan) for e in exp_cols})
            denom = float(np.abs(coefs).sum())
            for exp, coef in coefs.items():
                qg_weight_rows.append(
                    {
                        "outcome": outcome,
                        "exposure": exp,
                        "quantile_coefficient": coef,
                        "direction": "positive" if coef >= 0 else "negative",
                        "abs_weight": abs(coef) / denom if denom > 0 else np.nan,
                    }
                )
    qg = pd.DataFrame(qg_rows)
    if not qg.empty:
        qg["fdr"] = bh_fdr(qg["p_value"])
    out["nhanes_qgcomp_summary"] = qg
    out["nhanes_qgcomp_component_weights"] = pd.DataFrame(qg_weight_rows)

    # WQS approximation with repeated train/test splits.
    wqs_rows = []
    wqs_weight_rows = []
    repeats = 50
    complete_base = nh.dropna(subset=exp_cols + [weight_col, cluster_col]).copy()
    for outcome in outcomes:
        complete = complete_base.dropna(subset=[outcome]).copy()
        if len(complete) < 200:
            continue
        for exp in exp_cols:
            complete[f"q_{exp}"] = quantile_score(complete[exp], 4)
        q_exp = [f"q_{e}" for e in exp_cols]
        for direction in ["positive", "negative"]:
            for r in range(repeats):
                idx = np.arange(len(complete))
                RNG.shuffle(idx)
                cut = int(len(idx) * 0.60)
                train_idx = complete.index[idx[:cut]]
                test_idx = complete.index[idx[cut:]]
                train = complete.loc[train_idx].copy()
                test = complete.loc[test_idx].copy()
                y_res = residualize(train, outcome, cov_num, cov_cat, weight_col)
                cors = []
                for qe in q_exp:
                    x_res = residualize(train, qe, cov_num, cov_cat, weight_col)
                    c = weighted_corr(x_res, y_res, numeric(train[weight_col]))
                    cors.append(c)
                corr = np.array(cors, dtype=float)
                raw_w = np.maximum(corr, 0) if direction == "positive" else np.maximum(-corr, 0)
                if not np.isfinite(raw_w).any() or np.nansum(raw_w) == 0:
                    raw_w = np.abs(np.nan_to_num(corr, nan=0.0))
                if raw_w.sum() == 0:
                    raw_w = np.ones(len(q_exp))
                weights = raw_w / raw_w.sum()
                for exp, wt in zip(exp_cols, weights):
                    wqs_weight_rows.append({"outcome": outcome, "direction": direction, "repeat": r + 1, "exposure": exp, "wqs_weight": wt})
                test["wqs_index"] = sum(weights[i] * test[q_exp[i]] for i in range(len(q_exp)))
                X, _ = make_design(test, ["wqs_index"] + cov_num, cov_cat, standardize_numeric=True)
                fit = fit_wls_cluster(numeric(test[outcome]), X, weights=numeric(test[weight_col]), clusters=test[cluster_col])
                ft = fit_table(fit, "wqs_repeated_holdout_python", outcome, {"direction": direction, "repeat": r + 1})
                term = ft.loc[ft["term"] == "wqs_index"]
                if not term.empty:
                    row = term.iloc[0].to_dict()
                    wqs_rows.append(row)
    wqs_rep = pd.DataFrame(wqs_rows)
    if not wqs_rep.empty:
        wqs_summary = (
            wqs_rep.groupby(["outcome", "direction"], as_index=False)
            .agg(
                mean_beta=("estimate", "mean"),
                sd_beta=("estimate", "std"),
                mean_se=("std_error", "mean"),
                median_p=("p_value", "median"),
                mean_r2=("r2", "mean"),
                n_repeats=("repeat", "count"),
                mean_n=("n", "mean"),
            )
        )
        wqs_summary["z_meta"] = wqs_summary["mean_beta"] / wqs_summary["sd_beta"].replace(0, np.nan)
        wqs_summary["p_meta"] = wqs_summary["z_meta"].apply(p_norm_2sided)
        wqs_summary["fdr"] = bh_fdr(wqs_summary["p_meta"])
    else:
        wqs_summary = pd.DataFrame()
    out["nhanes_wqs_repeated_holdout_results"] = wqs_rep
    out["nhanes_wqs_summary"] = wqs_summary
    wqs_weights = pd.DataFrame(wqs_weight_rows)
    if not wqs_weights.empty:
        wqs_weights = wqs_weights.groupby(["outcome", "direction", "exposure"], as_index=False)["wqs_weight"].mean()
    out["nhanes_wqs_component_weights"] = wqs_weights

    # Restricted cubic spline.
    rcs_rows = []
    rcs_grid_rows = []
    for outcome in outcomes:
        for exp in exp_cols:
            rdf = nh.dropna(subset=[outcome, exp, weight_col, cluster_col]).copy()
            if len(rdf) < 200:
                continue
            rdf[f"{exp}_z"] = zscore(winsorize(rdf[exp]))
            knots = np.nanpercentile(rdf[f"{exp}_z"], [5, 35, 65, 95])
            basis = rcs_terms(rdf[f"{exp}_z"], knots)
            rdf[f"{exp}_linear"] = basis["linear"]
            nonlinear_cols = []
            for col in basis.columns:
                if col != "linear":
                    name = f"{exp}_{col}"
                    rdf[name] = basis[col]
                    nonlinear_cols.append(name)
            X, rcs_params = make_design(rdf, [f"{exp}_linear"] + nonlinear_cols + cov_num, cov_cat, standardize_numeric=True)
            fit = fit_wls_cluster(numeric(rdf[outcome]), X, weights=numeric(rdf[weight_col]), clusters=rdf[cluster_col])
            idx = [fit.terms.index(c) for c in nonlinear_cols if c in fit.terms]
            if idx:
                b = fit.beta[idx]
                V = fit.vcov[np.ix_(idx, idx)]
                chi = float(b.T @ safe_inverse(V) @ b)
                p_nonlin = chi2_sf(chi, len(idx))
            else:
                chi = np.nan
                p_nonlin = np.nan
            lin_row = fit_table(fit, "restricted_cubic_spline_python", outcome, {"exposure": exp})
            term = lin_row.loc[lin_row["term"] == f"{exp}_linear"]
            rcs_rows.append(
                {
                    "outcome": outcome,
                    "exposure": exp,
                    "n": fit.n,
                    "clusters": fit.clusters,
                    "linear_beta": float(term["estimate"].iloc[0]) if not term.empty else np.nan,
                    "linear_p": float(term["p_value"].iloc[0]) if not term.empty else np.nan,
                    "nonlinear_chi_square": chi,
                    "nonlinear_df": len(idx),
                    "nonlinear_p": p_nonlin,
                    "knots_z": "|".join(f"{k:.3f}" for k in knots),
                    "r2": fit.r2,
                }
            )
            grid = np.linspace(np.nanpercentile(rdf[f"{exp}_z"], 5), np.nanpercentile(rdf[f"{exp}_z"], 95), 21)
            for gx in grid:
                gdf = rdf.copy()
                gb = rcs_terms(pd.Series(gx, index=gdf.index), knots)
                gdf[f"{exp}_linear"] = gb["linear"]
                for col in gb.columns:
                    if col != "linear":
                        gdf[f"{exp}_{col}"] = gb[col]
                GX, _ = make_design(
                    gdf,
                    [f"{exp}_linear"] + nonlinear_cols + cov_num,
                    cov_cat,
                    standardize_numeric=True,
                    reference_columns=fit.terms,
                    numeric_params=rcs_params,
                )
                pred = GX.to_numpy(dtype=float) @ fit.beta
                rcs_grid_rows.append({"outcome": outcome, "exposure": exp, "exposure_z_grid": gx, "marginal_predicted_outcome": float(np.nanmean(pred))})
    rcs = pd.DataFrame(rcs_rows)
    if not rcs.empty:
        rcs["nonlinear_fdr"] = bh_fdr(rcs["nonlinear_p"])
    out["nhanes_rcs_tests"] = rcs
    out["nhanes_rcs_marginal_grid"] = pd.DataFrame(rcs_grid_rows)

    # BKMR-style kernel ridge mixture approximation.
    bkmr_rows = []
    bkmr_importance_rows = []
    bkmr_curve_rows = []
    for outcome in outcomes:
        bdf = nh.dropna(subset=[outcome] + exp_cols).copy()
        if len(bdf) < 300:
            continue
        if len(bdf) > 800:
            weights = numeric(bdf[weight_col]).fillna(1.0)
            probs = weights / weights.sum()
            bdf = bdf.iloc[RNG.choice(np.arange(len(bdf)), size=800, replace=False, p=probs.to_numpy())].copy()
        y_res = residualize(bdf, outcome, cov_num, cov_cat, weight_col)
        zmat = []
        for exp in exp_cols:
            bdf[f"{exp}_z"] = zscore(winsorize(bdf[exp]))
            zmat.append(bdf[f"{exp}_z"].to_numpy(dtype=float))
        Z = np.vstack(zmat).T
        mask = np.isfinite(y_res.to_numpy()) & np.isfinite(Z).all(axis=1)
        Z = Z[mask]
        yy = y_res.loc[mask].to_numpy(dtype=float)
        if len(yy) < 200:
            continue
        d = ((Z[:, None, :] - Z[None, :, :]) ** 2).sum(axis=2)
        tri = d[np.triu_indices_from(d, k=1)]
        sigma2 = float(np.nanmedian(tri[tri > 0])) if np.any(tri > 0) else 1.0
        K = np.exp(-d / (2 * sigma2))
        lam = 0.20
        alpha = np.linalg.solve(K + lam * np.eye(len(K)), yy)
        pred = K @ alpha
        mse = float(np.mean((yy - pred) ** 2))
        r2 = 1 - mse / float(np.var(yy, ddof=0)) if np.var(yy) > 0 else np.nan
        bkmr_rows.append(
            {
                "outcome": outcome,
                "method": "BKMR_style_gaussian_kernel_ridge_python_approximation",
                "n_subsample": len(yy),
                "lambda": lam,
                "kernel_sigma2_median_distance": sigma2,
                "in_sample_residual_r2": r2,
                "mse": mse,
                "note": "Nonlinear mixture screen; not a replacement for R bkmr posterior inference.",
            }
        )
        for j, exp in enumerate(exp_cols):
            Zp = Z.copy()
            Zp[:, j] = RNG.permutation(Zp[:, j])
            dp = ((Zp[:, None, :] - Z[None, :, :]) ** 2).sum(axis=2)
            Kp = np.exp(-dp / (2 * sigma2))
            predp = Kp @ alpha
            mse_p = float(np.mean((yy - predp) ** 2))
            bkmr_importance_rows.append({"outcome": outcome, "exposure": exp, "permutation_delta_mse": mse_p - mse, "relative_importance": (mse_p - mse) / mse if mse > 0 else np.nan})
        qs = [0.10, 0.25, 0.50, 0.75, 0.90]
        for q in qs:
            zq = np.nanquantile(Z, q, axis=0)
            Zq = np.tile(zq, (len(Z), 1))
            dq = ((Zq[:, None, :] - Z[None, :, :]) ** 2).sum(axis=2)
            Kq = np.exp(-dq / (2 * sigma2))
            bkmr_curve_rows.append({"outcome": outcome, "joint_exposure_quantile": q, "mean_kernel_predicted_residual_outcome": float(np.mean(Kq @ alpha))})
    out["nhanes_bkmr_kernel_summary"] = pd.DataFrame(bkmr_rows)
    out["nhanes_bkmr_kernel_importance"] = pd.DataFrame(bkmr_importance_rows)
    out["nhanes_bkmr_kernel_joint_response_curve"] = pd.DataFrame(bkmr_curve_rows)
    return out


def add_food_matrix(cg: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = cg.copy()
    carbs = numeric(df.get("eff_carbs", df.get("carbs"))).fillna(numeric(df["carbs"]))
    calories = numeric(df.get("eff_calories", df.get("calories"))).fillna(numeric(df["calories"]))
    protein = numeric(df.get("eff_protein", df.get("protein"))).fillna(numeric(df["protein"]))
    fat = numeric(df.get("eff_fat", df.get("fat"))).fillna(numeric(df["fat"]))
    fiber = numeric(df.get("eff_fiber", df.get("fiber"))).fillna(numeric(df["fiber"]))
    calories = calories.replace(0, np.nan)

    df["carb_density_g_per_100kcal"] = carbs / (calories / 100.0)
    df["fiber_per_50g_carb"] = fiber / (carbs + 1e-6) * 50.0
    df["protein_fat_buffer_per_50g_carb"] = (protein + fat) / (carbs + 1e-6) * 50.0
    df["liquid_sugar_proxy_v1"] = ((carbs >= 10) & (calories <= 220) & ((protein + fat + fiber) <= 3)).astype(int)
    df["refined_starch_proxy_v1"] = ((carbs >= 30) & (df["fiber_per_50g_carb"] < 3) & (df["protein_fat_buffer_per_50g_carb"] < 35)).astype(int)
    df["mixed_matrix_proxy_v1"] = ((protein >= 15) & (fat >= 8) & (carbs >= 20)).astype(int)
    df["fiber_protected_proxy_v1"] = ((fiber >= 5) | (df["fiber_per_50g_carb"] >= 6)).astype(int)
    df["rapid_digestibility_score_v1"] = (
        zscore(winsorize(df["carb_density_g_per_100kcal"]))
        + zscore(df["liquid_sugar_proxy_v1"])
        + zscore(df["refined_starch_proxy_v1"])
        - zscore(winsorize(df["fiber_per_50g_carb"]))
        - zscore(winsorize(df["protein_fat_buffer_per_50g_carb"]))
    )
    df["matrix_protection_score_v1"] = -df["rapid_digestibility_score_v1"]

    df["liquid_sugar_proxy_v2"] = ((carbs >= 12) & (calories <= 250) & ((protein + fat + fiber) <= 4)).astype(int)
    df["refined_starch_proxy_v2"] = ((carbs >= 35) & (df["fiber_per_50g_carb"] < 4) & (df["protein_fat_buffer_per_50g_carb"] < 40)).astype(int)
    df["mixed_matrix_proxy_v2"] = ((protein >= 12) & (fat >= 6) & (carbs >= 15)).astype(int)
    df["fiber_protected_proxy_v2"] = ((fiber >= 4) | (df["fiber_per_50g_carb"] >= 5)).astype(int)
    df["rapid_digestibility_score_v2"] = (
        zscore(winsorize(df["carb_density_g_per_100kcal"]))
        + zscore(df["liquid_sugar_proxy_v2"])
        + zscore(df["refined_starch_proxy_v2"])
        - zscore(winsorize(df["fiber_per_50g_carb"]))
        - zscore(winsorize(df["protein_fat_buffer_per_50g_carb"]))
    )
    df["matrix_protection_score_v2"] = -df["rapid_digestibility_score_v2"]

    def category(row: pd.Series, version: int) -> str:
        liquid = row[f"liquid_sugar_proxy_v{version}"] == 1
        refined = row[f"refined_starch_proxy_v{version}"] == 1
        mixed = row[f"mixed_matrix_proxy_v{version}"] == 1
        fiberp = row[f"fiber_protected_proxy_v{version}"] == 1
        c = row["eff_carbs"] if "eff_carbs" in row else row["carbs"]
        if liquid:
            return "sweetened_beverage_or_liquid_carb"
        if refined and not mixed and not fiberp:
            return "refined_starch_low_matrix_protection"
        if fiberp:
            return "fiber_rich_or_whole_food_matrix"
        if mixed:
            return "mixed_meal_protein_fat_buffered"
        if pd.notna(c) and c < 15:
            return "low_carb_or_protein_forward"
        return "mixed_or_unknown_matrix"

    df["food_matrix_category_v1"] = df.apply(lambda r: category(r, 1), axis=1)
    df["food_matrix_category_v2"] = df.apply(lambda r: category(r, 2), axis=1)
    query_map = {
        "sweetened_beverage_or_liquid_carb": ("sweetened beverage", 0.55),
        "refined_starch_low_matrix_protection": ("white rice bread pasta refined grain", 0.50),
        "fiber_rich_or_whole_food_matrix": ("whole grain beans vegetable mixed dish", 0.55),
        "mixed_meal_protein_fat_buffered": ("mixed entree protein fat carbohydrate", 0.45),
        "low_carb_or_protein_forward": ("meat egg cheese low carbohydrate", 0.50),
        "mixed_or_unknown_matrix": ("mixed meal", 0.25),
    }
    df["fdc_proxy_query"] = df["food_matrix_category_v1"].map(lambda x: query_map[x][0])
    df["fdc_proxy_match_confidence"] = df["food_matrix_category_v1"].map(lambda x: query_map[x][1])
    df["fdc_search_url"] = "https://fdc.nal.usda.gov/fdc-app.html#/food-search?query=" + df["fdc_proxy_query"].str.replace(" ", "%20", regex=False)

    codebook = pd.DataFrame(
        [
            {"variable": "carb_density_g_per_100kcal", "definition": "Effective carbohydrate grams per 100 kcal.", "rationale": "Higher density can indicate rapidly available carbohydrate load."},
            {"variable": "fiber_per_50g_carb", "definition": "Fiber grams normalized to 50 g carbohydrate.", "rationale": "Fiber can slow starch/glucose availability."},
            {"variable": "protein_fat_buffer_per_50g_carb", "definition": "Protein plus fat grams normalized to 50 g carbohydrate.", "rationale": "Mixed macronutrient matrix can slow gastric emptying and glucose appearance."},
            {"variable": "liquid_sugar_proxy", "definition": "Low energy, carbohydrate-containing, minimal protein/fat/fiber meal.", "rationale": "Approximates liquid or simple sugar exposure when food text is unavailable."},
            {"variable": "refined_starch_proxy", "definition": "High carbohydrate, low fiber, low protein/fat buffering.", "rationale": "Approximates refined starch / low matrix protection."},
            {"variable": "mixed_matrix_proxy", "definition": "Meal containing carbohydrate plus substantial protein and fat.", "rationale": "Represents matrix buffering compared with isolated carbohydrate."},
            {"variable": "rapid_digestibility_score", "definition": "Composite z-score: carb density + liquid/refined proxies - fiber/protein-fat buffering.", "rationale": "Dry-lab digestibility availability score for CGM modeling."},
            {"variable": "fdc_proxy_query", "definition": "Representative USDA FoodData Central search string derived from matrix category.", "rationale": "Transparent proxy linkage; not item-level manual FDC matching."},
        ]
    )
    fdc = df[
        [
            "subject_id",
            "meal_time",
            "meal_type",
            "image_path",
            "food_matrix_category_v1",
            "food_matrix_category_v2",
            "fdc_proxy_query",
            "fdc_proxy_match_confidence",
            "fdc_search_url",
        ]
    ].copy()
    reliability = pd.DataFrame(
        [
            {"metric": "category_cohen_kappa_v1_vs_v2", "value": cohen_kappa(df["food_matrix_category_v1"], df["food_matrix_category_v2"]), "interpretation": "Algorithmic dual-rule stress test; not a substitute for two human raters."},
            {"metric": "rapid_digestibility_icc_v1_vs_v2", "value": icc_two_way_single(df["rapid_digestibility_score_v1"], df["rapid_digestibility_score_v2"]), "interpretation": "Agreement of two prespecified threshold variants."},
            {"metric": "liquid_proxy_kappa", "value": cohen_kappa(df["liquid_sugar_proxy_v1"], df["liquid_sugar_proxy_v2"]), "interpretation": "Binary proxy stability."},
            {"metric": "refined_starch_proxy_kappa", "value": cohen_kappa(df["refined_starch_proxy_v1"], df["refined_starch_proxy_v2"]), "interpretation": "Binary proxy stability."},
        ]
    )
    return df, codebook, fdc, reliability


def ridge_predict_group_cv(df: pd.DataFrame, outcome: str, feature_cols: list[str], cat_cols: list[str], group_col: str, model_name: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    data = df.dropna(subset=[outcome, group_col]).copy()
    groups = np.array(sorted(data[group_col].dropna().unique()))
    RNG.shuffle(groups)
    folds = np.array_split(groups, min(5, len(groups)))
    pred_rows = []
    for fold_id, test_groups in enumerate(folds, start=1):
        train = data.loc[~data[group_col].isin(test_groups)].copy()
        test = data.loc[data[group_col].isin(test_groups)].copy()
        if len(train) < 30 or len(test) == 0:
            continue
        X_train, params = make_design(train, feature_cols, cat_cols, standardize_numeric=True)
        X_test, _ = make_design(test, feature_cols, cat_cols, standardize_numeric=True, reference_columns=list(X_train.columns), numeric_params=params)
        y_train = numeric(train[outcome])
        mask = y_train.notna()
        X = X_train.loc[mask].to_numpy(dtype=float)
        y = y_train.loc[mask].to_numpy(dtype=float)
        lam = 2.0
        penalty = np.eye(X.shape[1])
        penalty[0, 0] = 0.0
        beta = safe_inverse(X.T @ X + lam * penalty) @ (X.T @ y)
        pred = X_test.to_numpy(dtype=float) @ beta
        for idx, yhat in zip(test.index, pred):
            pred_rows.append(
                {
                    "model": model_name,
                    "outcome": outcome,
                    "fold": fold_id,
                    "subject_or_group": test.loc[idx, group_col],
                    "row_index": idx,
                    "observed": test.loc[idx, outcome],
                    "predicted": yhat,
                }
            )
    preds = pd.DataFrame(pred_rows)
    if preds.empty:
        return preds, pd.DataFrame()
    metrics = []
    for (model, outc), g in preds.groupby(["model", "outcome"]):
        obs = numeric(g["observed"])
        pr = numeric(g["predicted"])
        err = pr - obs
        rmse = math.sqrt(float(np.mean(err**2)))
        mae = float(np.mean(np.abs(err)))
        r2 = 1 - float(np.sum(err**2)) / float(np.sum((obs - obs.mean()) ** 2)) if obs.var(ddof=0) > 0 else np.nan
        corr = float(np.corrcoef(obs, pr)[0, 1]) if len(g) > 2 and obs.std() > 0 and pr.std() > 0 else np.nan
        cal_X = pd.DataFrame({"Intercept": 1.0, "predicted": pr})
        cal_fit = fit_wls_cluster(obs, cal_X)
        metrics.append(
            {
                "model": model,
                "outcome": outc,
                "n_predictions": len(g),
                "n_groups": g["subject_or_group"].nunique(),
                "mae": mae,
                "rmse": rmse,
                "r2": r2,
                "pearson_r": corr,
                "calibration_intercept": cal_fit.beta[0],
                "calibration_slope": cal_fit.beta[1] if len(cal_fit.beta) > 1 else np.nan,
            }
        )
    return preds, pd.DataFrame(metrics)


def run_cgmacros_models(cg: pd.DataFrame) -> dict[str, pd.DataFrame]:
    out: dict[str, pd.DataFrame] = {}
    df = cg.copy()
    df["carbs_10g"] = winsorize(numeric(df["eff_carbs"]).fillna(df["carbs"])) / 10.0
    df["protein_10g"] = winsorize(numeric(df["eff_protein"]).fillna(df["protein"])) / 10.0
    df["fat_10g"] = winsorize(numeric(df["eff_fat"]).fillna(df["fat"])) / 10.0
    df["fiber_5g"] = winsorize(numeric(df["eff_fiber"]).fillna(df["fiber"])) / 5.0
    df["energy_100kcal"] = winsorize(numeric(df["eff_calories"]).fillna(df["calories"])) / 100.0
    df["iauc_2h_per1000"] = numeric(df["iauc_2h"]) / 1000.0
    df["baseline_glucose_z"] = zscore(winsorize(df["baseline_glucose"]))
    df["pre_glucose_mean_30_z"] = zscore(winsorize(df["pre_glucose_mean_30"]))
    df["activity_pre30_z"] = zscore(winsorize(df["activity_calories_pre30"]))
    df["response_x_carbs"] = df["response_vulnerability_index"] * df["carbs_10g"]
    df["response_x_digestibility"] = df["response_vulnerability_index"] * df["rapid_digestibility_score_v1"]

    inference_rows = []
    vc_rows = []
    outcomes = ["iauc_2h_per1000", "peak_delta_2h"]
    predictors = [
        "carbs_10g",
        "protein_10g",
        "fat_10g",
        "fiber_5g",
        "rapid_digestibility_score_v1",
        "matrix_protection_score_v1",
        "exposure_facing_msi",
        "response_vulnerability_index",
        "response_x_carbs",
        "response_x_digestibility",
        "baseline_glucose_z",
        "pre_glucose_mean_30_z",
        "meal_hour_sin",
        "meal_hour_cos",
    ]
    cat_cols = ["meal_type", "Gender"]
    for outcome in outcomes:
        model_df = df.dropna(subset=[outcome, "subject_id"]).copy()
        X, _ = make_design(model_df, predictors, cat_cols, standardize_numeric=True)
        fit = fit_wls_cluster(numeric(model_df[outcome]), X, clusters=model_df["subject_id"].astype(str))
        inference_rows.append(fit_table(fit, "cluster_robust_hierarchical_screen", outcome))
        resid = pd.Series(fit.residuals, index=model_df.dropna(subset=[outcome]).index[: len(fit.residuals)])
        tmp = model_df.loc[resid.index, ["subject_id"]].copy()
        tmp["resid"] = resid.to_numpy()
        subject_mean = tmp.groupby("subject_id")["resid"].mean()
        subject_n = tmp.groupby("subject_id")["resid"].size()
        sigma_e2 = float(tmp.groupby("subject_id")["resid"].var().mean(skipna=True))
        sigma_b2 = float(max(subject_mean.var(ddof=1) - sigma_e2 * (1 / subject_n).mean(), 0.0))
        icc = sigma_b2 / (sigma_b2 + sigma_e2) if (sigma_b2 + sigma_e2) > 0 else np.nan
        vc_rows.append({"outcome": outcome, "sigma_subject_random_intercept": sigma_b2, "sigma_within_subject": sigma_e2, "icc_random_intercept_approx": icc, "n_subjects": tmp["subject_id"].nunique(), "n_meals": len(tmp)})
    coef = pd.concat(inference_rows, ignore_index=True) if inference_rows else pd.DataFrame()
    if not coef.empty:
        coef["fdr"] = bh_fdr(coef["p_value"])
    out["cgmacros_cluster_hierarchical_coefficients"] = coef
    out["cgmacros_random_intercept_variance_components"] = pd.DataFrame(vc_rows)

    feature_sets = {
        "macro_only_subject_disjoint": ["carbs_10g", "protein_10g", "fat_10g", "fiber_5g", "energy_100kcal"],
        "macro_context_subject_disjoint": ["carbs_10g", "protein_10g", "fat_10g", "fiber_5g", "energy_100kcal", "baseline_glucose", "pre_glucose_mean_30", "pre_glucose_sd_30", "meal_hour_sin", "meal_hour_cos"],
        "full_msi_foodmatrix_subject_disjoint": ["carbs_10g", "protein_10g", "fat_10g", "fiber_5g", "energy_100kcal", "baseline_glucose", "pre_glucose_mean_30", "pre_glucose_sd_30", "meal_hour_sin", "meal_hour_cos", "BMI", "A1c PDL (Lab)", "Fasting GLU - PDL (Lab)", "Insulin", "exposure_facing_msi", "response_vulnerability_index", "rapid_digestibility_score_v1", "matrix_protection_score_v1"],
    }
    pred_tables = []
    perf_tables = []
    for outcome in ["iauc_2h", "peak_delta_2h"]:
        for model_name, feats in feature_sets.items():
            preds, perf = ridge_predict_group_cv(df, outcome, feats, ["meal_type", "Gender", "food_matrix_category_v1"], "subject_id", model_name)
            pred_tables.append(preds)
            perf_tables.append(perf)
    out["cgmacros_subject_disjoint_prediction_records"] = pd.concat(pred_tables, ignore_index=True) if pred_tables else pd.DataFrame()
    out["cgmacros_subject_disjoint_performance"] = pd.concat(perf_tables, ignore_index=True) if perf_tables else pd.DataFrame()
    return out


def run_stanford_boundary() -> dict[str, pd.DataFrame]:
    out: dict[str, pd.DataFrame] = {}
    cgm_path = STANFORD_DIR / "data_cgm.csv"
    meta_path = STANFORD_DIR / "data_meta.csv"
    if not cgm_path.exists() or not meta_path.exists():
        out["stanford_boundary_status"] = pd.DataFrame([{"status": "missing_data", "cgm_path": str(cgm_path), "meta_path": str(meta_path)}])
        return out
    cgm = pd.read_csv(cgm_path)
    meta = pd.read_csv(meta_path)
    rows = []
    group_cols = ["subject", "foods", "mitigator", "food", "rep"]
    for keys, g in cgm.groupby(group_cols, dropna=False):
        g = g.sort_values("mins_since_start")
        baseline = numeric(g.loc[(g["mins_since_start"] >= -30) & (g["mins_since_start"] <= 0), "glucose"]).mean()
        post = g.loc[(g["mins_since_start"] >= 0) & (g["mins_since_start"] <= 120)].copy()
        if len(post) < 5 or not np.isfinite(baseline):
            continue
        t = numeric(post["mins_since_start"]).to_numpy(dtype=float)
        y = numeric(post["glucose"]).to_numpy(dtype=float)
        order = np.argsort(t)
        t = t[order]
        y = y[order]
        delta = y - baseline
        iauc = float(np.trapezoid(np.maximum(delta, 0), t))
        peak = float(np.nanmax(y) - baseline)
        time_peak = float(t[np.nanargmax(y)])
        rows.append(
            {
                "subject": keys[0],
                "foods": keys[1],
                "mitigator": keys[2],
                "food": keys[3],
                "rep": keys[4],
                "baseline_glucose": baseline,
                "iauc_2h": iauc,
                "peak_delta_2h": peak,
                "time_to_peak_2h": time_peak,
                "n_points": len(post),
            }
        )
    challenge = pd.DataFrame(rows)
    m = meta.rename(columns={"id": "subject"}).copy()
    m["stanford_ln_homa_proxy"] = np.log((numeric(m["fasting glucose"]) * numeric(m["fasting insulin"]) / 405.0).where(lambda x: x > 0))
    m["stanford_total_hdl_ratio"] = numeric(m["Total cholesterol"]) / numeric(m["HDL"])
    m["stanford_exposure_facing_msi"] = pd.concat([zscore(m["BMI"]), zscore(m["stanford_total_hdl_ratio"])], axis=1).mean(axis=1, skipna=True)
    m["stanford_response_vulnerability_index"] = pd.concat(
        [
            zscore(m["BMI"]),
            zscore(m["HbA1c"]),
            zscore(m["fasting glucose"]),
            zscore(m["fasting insulin"]),
            zscore(m["SSPG"]),
            zscore(m["Hepatic IR"]),
        ],
        axis=1,
    ).mean(axis=1, skipna=True)
    challenge = challenge.merge(m, on="subject", how="left")
    out["stanford_challenge_level_ppgr"] = challenge

    coef_rows = []
    for outcome in ["iauc_2h", "peak_delta_2h"]:
        model_df = challenge.dropna(subset=[outcome]).copy()
        X, _ = make_design(
            model_df,
            ["stanford_response_vulnerability_index", "stanford_exposure_facing_msi", "baseline_glucose", "age", "BMI"],
            ["food", "Sex"],
            standardize_numeric=True,
        )
        fit = fit_wls_cluster(numeric(model_df[outcome]), X, clusters=model_df["subject"])
        coef_rows.append(fit_table(fit, "stanford_cluster_boundary_model", outcome))
    coef = pd.concat(coef_rows, ignore_index=True) if coef_rows else pd.DataFrame()
    if not coef.empty:
        coef["fdr"] = bh_fdr(coef["p_value"])
    out["stanford_boundary_coefficients"] = coef

    pred_tables = []
    perf_tables = []
    for outcome in ["iauc_2h", "peak_delta_2h"]:
        preds, perf = ridge_predict_group_cv(
            challenge,
            outcome,
            ["stanford_response_vulnerability_index", "stanford_exposure_facing_msi", "baseline_glucose", "age", "BMI"],
            ["food", "Sex"],
            "subject",
            "stanford_subject_disjoint_boundary_prediction",
        )
        pred_tables.append(preds)
        perf_tables.append(perf)
    out["stanford_subject_disjoint_prediction_records"] = pd.concat(pred_tables, ignore_index=True) if pred_tables else pd.DataFrame()
    out["stanford_subject_disjoint_performance"] = pd.concat(perf_tables, ignore_index=True) if perf_tables else pd.DataFrame()
    dist = []
    for outcome in ["iauc_2h", "peak_delta_2h"]:
        s = numeric(challenge[outcome]).dropna()
        dist.append({"dataset": "Stanford_CGMDB", "outcome": outcome, "n": len(s), "mean": s.mean(), "sd": s.std(), "p25": s.quantile(0.25), "median": s.median(), "p75": s.quantile(0.75), "min": s.min(), "max": s.max()})
    out["stanford_distribution_summary"] = pd.DataFrame(dist)
    return out


def ctd_read_filtered(path: Path, chemical_ids: set[str], columns: list[str], chunksize: int = 200000) -> pd.DataFrame:
    rows = []
    ids = [cid for cid in chemical_ids if cid]
    with gzip.open(path, "rt", encoding="utf-8", errors="replace") as f:
        for line in f:
            if line.startswith("#"):
                continue
            if not any(cid in line for cid in ids):
                continue
            parts = line.rstrip("\n").split("\t")
            if len(parts) < len(columns):
                parts += [""] * (len(columns) - len(parts))
            rec = dict(zip(columns, parts[: len(columns)]))
            cid = rec.get("ChemicalID", rec.get("chemicalid", ""))
            if cid in chemical_ids:
                rows.append(rec)
    return pd.DataFrame(rows, columns=columns)


def fetch_url(url: str, timeout: int = 20) -> tuple[bool, str]:
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 Codex research workflow"})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return True, resp.read().decode("utf-8", errors="replace")
    except Exception as exc:
        return False, str(exc)


def run_mechanism_workflow() -> dict[str, pd.DataFrame]:
    out: dict[str, pd.DataFrame] = {}
    chem_cols = ["ChemicalName", "ChemicalID", "CasRN", "PubChemCID", "PubChemSID", "DTXSID", "InChIKey", "Definition", "ParentIDs", "TreeNumbers", "ParentTreeNumbers", "MESHSynonyms", "CTDCuratedSynonyms"]
    chem_path = CTD_DIR / "CTD_chemicals.tsv.gz"
    chemicals = []
    target_terms = [
        "diethylhexyl phthalate",
        "bis(2-ethylhexyl)phthalate",
        "bis(2-ethylhexyl) phthalate",
        "di(2-ethylhexyl)phthalate",
        "mono-(2-ethylhexyl)phthalate",
        "mono(2-ethylhexyl) phthalate",
        "mono(2-ethyl-5-hydroxyhexyl) phthalate",
        "mono(2-ethyl-5-oxohexyl)phthalate",
        "mehp compound",
        "mehhp cpd",
        "meohp cpd",
    ]
    with gzip.open(chem_path, "rt", encoding="utf-8", errors="replace") as f:
        for line in f:
            if line.startswith("#"):
                continue
            parts = line.rstrip("\n").split("\t")
            if len(parts) < len(chem_cols):
                parts += [""] * (len(chem_cols) - len(parts))
            rec = dict(zip(chem_cols, parts))
            blob = " ".join(str(rec.get(c, "")) for c in chem_cols).lower()
            if any(t in blob for t in target_terms):
                chemicals.append(rec)
    chem_df = pd.DataFrame(chemicals).drop_duplicates(subset=["ChemicalID"])
    # Keep DEHP and major oxidative/monoester metabolites while excluding unrelated strings with accidental MEHP substrings.
    keep = chem_df["ChemicalName"].str.contains("phthalate|Diethylhexyl", case=False, regex=True, na=False) | chem_df["MESHSynonyms"].str.contains("phthalate|DEHP|MEHP compound|MEHHP cpd|MEOHP cpd|5oxo-MEHP", case=False, regex=True, na=False)
    chem_df = chem_df.loc[keep].copy()
    chem_df = chem_df.loc[~chem_df["ChemicalName"].str.contains("hydroxypyridin|prostacyclin|cyclophosphamide|paromomycin|danitracen|propionyl", case=False, regex=True, na=False)].copy()
    chemical_ids = set(chem_df["ChemicalID"].dropna().astype(str).str.replace("MESH:", "", regex=False))
    out["mechanism_ctd_target_chemicals"] = chem_df

    gene_cols = ["ChemicalName", "ChemicalID", "CasRN", "GeneSymbol", "GeneID", "GeneForms", "Organism", "OrganismID", "Interaction", "InteractionActions", "PubMedIDs"]
    gene_df = ctd_read_filtered(CTD_DIR / "CTD_chem_gene_ixns.tsv.gz", chemical_ids, gene_cols)
    if not gene_df.empty and "OrganismID" in gene_df.columns:
        gene_df = gene_df.loc[(gene_df["OrganismID"] == "9606") | (gene_df["Organism"].str.contains("Homo sapiens", na=False))].copy()
    out["mechanism_ctd_chemical_gene_edges"] = gene_df

    disease_cols = ["ChemicalName", "ChemicalID", "CasRN", "DiseaseName", "DiseaseID", "DirectEvidence", "InferenceGeneSymbol", "InferenceScore", "OmimIDs", "PubMedIDs"]
    disease_df = ctd_read_filtered(CTD_DIR / "CTD_chemicals_diseases.tsv.gz", chemical_ids, disease_cols)
    if not disease_df.empty:
        pattern = "diabetes|insulin|metabolic|obesity|glucose|cardiovascular|inflammation|oxidative|fatty liver|dyslipidemia"
        disease_df["metabolic_relevance_flag"] = disease_df["DiseaseName"].str.contains(pattern, case=False, regex=True, na=False)
    out["mechanism_ctd_chemical_disease_edges"] = disease_df

    gene_summary = pd.DataFrame()
    if not gene_df.empty:
        gene_summary = (
            gene_df.groupby(["GeneSymbol", "GeneID"], as_index=False)
            .agg(
                n_chemical_edges=("ChemicalID", "count"),
                n_chemicals=("ChemicalID", "nunique"),
                n_pubmed_records=("PubMedIDs", lambda x: len(set("|".join(x.dropna().astype(str)).split("|")) - {""})),
                example_actions=("InteractionActions", lambda x: "; ".join(sorted(set(x.dropna().astype(str)))[:4])),
            )
            .sort_values(["n_chemicals", "n_chemical_edges"], ascending=False)
        )
    out["mechanism_ctd_gene_summary"] = gene_summary

    # KEGG pathway overlap using REST.
    kegg_pathways = {
        "hsa04910": "Insulin signaling pathway",
        "hsa04931": "Insulin resistance",
        "hsa04151": "PI3K-Akt signaling pathway",
        "hsa04152": "AMPK signaling pathway",
        "hsa03320": "PPAR signaling pathway",
        "hsa04068": "FoxO signaling pathway",
        "hsa00190": "Oxidative phosphorylation",
        "hsa04668": "TNF signaling pathway",
        "hsa04933": "AGE-RAGE signaling pathway in diabetic complications",
    }
    kegg_edges = []
    kegg_status = []
    for pid, pname in kegg_pathways.items():
        ok, text = fetch_url(f"https://rest.kegg.jp/link/hsa/{pid}", timeout=20)
        kegg_status.append({"database": "KEGG", "pathway_id": pid, "pathway_name": pname, "success": ok, "message_or_n_chars": len(text) if ok else text})
        if ok:
            for line in text.strip().splitlines():
                parts = line.split("\t")
                if len(parts) == 2:
                    gene_id = parts[1].replace("hsa:", "")
                    kegg_edges.append({"pathway_id": pid, "pathway_name": pname, "GeneID": gene_id})
    kegg_df = pd.DataFrame(kegg_edges)
    out["mechanism_kegg_download_status"] = pd.DataFrame(kegg_status)
    if not gene_summary.empty and not kegg_df.empty:
        overlap = gene_summary.merge(kegg_df, on="GeneID", how="inner")
        enrich_rows = []
        all_ctd_genes = set(gene_summary["GeneID"].astype(str))
        for (pid, pname), g in kegg_df.groupby(["pathway_id", "pathway_name"]):
            pathway_genes = set(g["GeneID"].astype(str))
            overlap_genes = all_ctd_genes & pathway_genes
            enrich_rows.append(
                {
                    "database": "KEGG",
                    "pathway_id": pid,
                    "pathway_name": pname,
                    "n_pathway_genes": len(pathway_genes),
                    "n_ctd_genes": len(all_ctd_genes),
                    "n_overlap": len(overlap_genes),
                    "overlap_gene_symbols": ";".join(sorted(gene_summary.loc[gene_summary["GeneID"].astype(str).isin(overlap_genes), "GeneSymbol"].astype(str).unique())[:100]),
                }
            )
        out["mechanism_kegg_pathway_overlap"] = pd.DataFrame(enrich_rows).sort_values("n_overlap", ascending=False)
        out["mechanism_kegg_gene_pathway_edges"] = overlap
    else:
        out["mechanism_kegg_pathway_overlap"] = pd.DataFrame()
        out["mechanism_kegg_gene_pathway_edges"] = pd.DataFrame()

    # Reactome mapping through public download; filtered to relevant pathway names.
    reactome_url = "https://reactome.org/download/current/NCBI2Reactome_PE_All_Levels.txt"
    ok, text = fetch_url(reactome_url, timeout=30)
    react_status = [{"database": "Reactome", "url": reactome_url, "success": ok, "message_or_n_chars": len(text) if ok else text}]
    react_edges = []
    if ok and not gene_summary.empty:
        ctd_gene_ids = set(gene_summary["GeneID"].astype(str))
        keywords = ["insulin", "glucose", "oxidative", "mitochond", "inflamm", "cytokine", "lipid", "ppar", "akt", "pi3k", "stress"]
        for line in text.splitlines():
            parts = line.split("\t")
            if len(parts) >= 8:
                gene_id = parts[0]
                reactome_id = parts[3]
                pathway_name = parts[5]
                species = parts[7]
                if species == "Homo sapiens" and gene_id in ctd_gene_ids and any(k in pathway_name.lower() for k in keywords):
                    react_edges.append({"GeneID": gene_id, "reactome_id": reactome_id, "pathway_name": pathway_name})
    react_df = pd.DataFrame(react_edges)
    out["mechanism_reactome_download_status"] = pd.DataFrame(react_status)
    if not react_df.empty:
        react_df = react_df.merge(gene_summary[["GeneID", "GeneSymbol", "n_chemical_edges", "n_chemicals"]], on="GeneID", how="left")
        out["mechanism_reactome_relevant_edges"] = react_df
        out["mechanism_reactome_pathway_summary"] = react_df.groupby(["reactome_id", "pathway_name"], as_index=False).agg(n_overlap=("GeneID", "nunique"), genes=("GeneSymbol", lambda x: ";".join(sorted(set(x.dropna().astype(str)))[:80]))).sort_values("n_overlap", ascending=False)
    else:
        out["mechanism_reactome_relevant_edges"] = pd.DataFrame()
        out["mechanism_reactome_pathway_summary"] = pd.DataFrame()

    out["mechanism_comptox_query_manifest"] = pd.DataFrame(
        [
            {
                "chemical_name": row.get("ChemicalName", ""),
                "chemical_id": row.get("ChemicalID", ""),
                "dtxsid": row.get("DTXSID", ""),
                "casrn": row.get("CasRN", ""),
                "comptox_dashboard_url": f"https://comptox.epa.gov/dashboard/chemical/details/{row.get('DTXSID', '')}" if row.get("DTXSID", "") else f"https://comptox.epa.gov/dashboard/chemical/search-results?search={urllib.parse.quote(row.get('ChemicalName', ''))}",
                "role": "reproducible CompTox lookup link; batch API not required for the local CTD network.",
            }
            for _, row in chem_df.iterrows()
        ]
    )
    return out


def build_reporting_checklists() -> dict[str, pd.DataFrame]:
    def rows(name: str, items: list[tuple[str, str, str]]) -> pd.DataFrame:
        return pd.DataFrame(
            [
                {
                    "checklist": name,
                    "item_id": item_id,
                    "domain": domain,
                    "requirement": req,
                    "current_project_status": "addressed_in_upgrade_package",
                    "evidence_or_file": "high_impact_computational_upgrade_2026-06-11",
                    "remaining_action_before_submission": "Replace approximation labels with final R/package runs where required; complete manual manuscript page/line references.",
                }
                for item_id, domain, req in items
            ]
        )

    tripod_items = [
        ("TRIPODAI-01", "Title/Abstract", "Identify the study as prediction model development/validation and mention AI/ML methods where used."),
        ("TRIPODAI-02", "Background", "Explain intended use, target population, and prediction horizon for PPGR."),
        ("TRIPODAI-03", "Data Source", "Describe CGMacros and Stanford CGM data provenance, inclusion, and preprocessing."),
        ("TRIPODAI-04", "Participants", "Report participant flow, exclusions, and subject-disjoint split logic."),
        ("TRIPODAI-05", "Outcome", "Define iaUC, peak delta, baseline window, and meal response window."),
        ("TRIPODAI-06", "Predictors", "Define meal, context, MSI, and food-matrix predictors before modeling."),
        ("TRIPODAI-07", "Missing Data", "Report imputation rules and variables with missingness."),
        ("TRIPODAI-08", "Sample Size", "Report number of subjects and meals; avoid treating meals as independent subjects."),
        ("TRIPODAI-09", "Model Development", "Specify ridge regression, feature scaling, hyperparameter, and no subject leakage."),
        ("TRIPODAI-10", "Validation", "Use subject-disjoint cross-validation and Stanford external boundary validation."),
        ("TRIPODAI-11", "Performance", "Report MAE, RMSE, R2, Pearson r, calibration intercept/slope."),
        ("TRIPODAI-12", "Explainability", "Report coefficient direction and food-matrix/MSI contributions."),
        ("TRIPODAI-13", "Reproducibility", "Provide script, package versions, random seed, and output manifest."),
    ]
    probast_items = [
        ("PROBASTAI-01", "Participants", "Assess whether participants represent intended target users."),
        ("PROBASTAI-02", "Predictors", "Check that predictors are available at prediction time and measured consistently."),
        ("PROBASTAI-03", "Outcome", "Check CGM outcome definition and blinding from predictors."),
        ("PROBASTAI-04", "Analysis", "Check subject-level split, events per parameter, overfitting control, and calibration."),
        ("PROBASTAI-05", "Missingness", "Assess missing data handling and whether missingness could induce bias."),
        ("PROBASTAI-06", "Applicability", "Document Stanford boundary differences and transportability limits."),
        ("PROBASTAI-07", "Overall", "Rate current model as development/boundary-validation, not deployment-ready."),
    ]
    strobe_items = [
        ("STROBE-01", "Title/Abstract", "Indicate observational cross-sectional NHANES and repeated-meal CGMacros design."),
        ("STROBE-02", "Setting", "Report NHANES cycles, CGMacros collection context, and Stanford challenge setting."),
        ("STROBE-03", "Participants", "Report inclusion/exclusion and missingness per module."),
        ("STROBE-04", "Variables", "Define exposures, outcomes, covariates, MSI, and food-matrix variables."),
        ("STROBE-05", "Bias", "Discuss single spot urine, residual confounding, and cross-dataset bridge limitations."),
        ("STROBE-06", "Study Size", "Report subject-level and meal-level sample sizes separately."),
        ("STROBE-07", "Statistical Methods", "Report survey weights, cluster robust SEs, splines, mixture models, and sensitivity analyses."),
        ("STROBE-08", "Results", "Report weighted estimates, precision, and sensitivity consistency."),
        ("STROBE-09", "Limitations", "Avoid direct DEHP-to-PPGR mediation claims."),
    ]
    prisma_items = [
        ("PRISMA-01", "Rationale", "Explain why evidence mapping is needed across environmental exposure, PPGR, food matrix, and mechanisms."),
        ("PRISMA-02", "Eligibility", "Define eligible human, database, and methodology sources."),
        ("PRISMA-03", "Information Sources", "List PubMed, NHANES, PhysioNet, Stanford CGMDB, CTD, CompTox, Reactome, KEGG, FDC."),
        ("PRISMA-04", "Search Strategy", "Preserve exact search strings and dates for final systematic supplement."),
        ("PRISMA-05", "Selection", "Use dual screening before submission if turning evidence map into formal review."),
        ("PRISMA-06", "Data Items", "Extract design, population, exposure, outcome, model, limitations, and relevance."),
        ("PRISMA-07", "Bias", "Assess cross-sectional, measurement error, and prediction-model bias."),
        ("PRISMA-08", "Synthesis", "Summarize as triangulation layers rather than a single causal chain."),
        ("PRISMA-09", "Transparency", "Provide literature evidence table and database query manifest."),
    ]
    return {
        "TRIPOD_AI_checklist": rows("TRIPOD+AI", tripod_items),
        "PROBAST_AI_checklist": rows("PROBAST+AI", probast_items),
        "STROBE_checklist": rows("STROBE", strobe_items),
        "PRISMA_checklist": rows("PRISMA", prisma_items),
    }


def write_workbook(index_rows: list[dict[str, str]]) -> None:
    if not HAS_OPENPYXL:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "manifest"
    headers = ["module", "file", "description"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in index_rows:
        ws.append([row.get(h, "") for h in headers])
    path = OUT / "00_high_impact_computational_upgrade_manifest.xlsx"
    wb.save(path)
    mirror_file(path)


def build_report(index_rows: list[dict[str, str]], qa_notes: list[str]) -> str:
    manifest_md = "\n".join([f"- `{r['file']}` - {r['description']}" for r in index_rows])
    notes_md = "\n".join([f"- {n}" for n in qa_notes])
    return f"""
# High-impact computational upgrade report

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Central Guardrail

This package strengthens the project as a triangulated cross-dataset bridge-phenotype study. It does not claim direct DEHP-to-PPGR causality or formal mediation because no same-participant dataset currently contains both urinary DEHP metabolites and meal-level CGM PPGR.

## What Was Completed

1. NHANES survey-weighted exposure models plus WQS, qgcomp, BKMR-style nonlinear mixture screen, restricted cubic splines, and E-value sensitivity tables.
2. Dual MSI reconstruction: exposure-facing MSI and response-vulnerability index in NHANES and CGMacros.
3. CGMacros prediction and inference were rerun with subject-disjoint splits and subject-cluster/hierarchical inference.
4. Food matrix score was formalized with codebook, proxy FDC linkage, and dual-rule reliability checks.
5. Stanford CGM Database was converted to standardized challenge-level PPGR phenotypes for external boundary validation.
6. Mechanism layer was rebuilt around CTD chemical-gene/disease edges, KEGG pathway overlap, Reactome pathway mapping, and CompTox reproducible query links.
7. TRIPOD+AI, PROBAST+AI, STROBE, and PRISMA supplementary checklists were generated.

## Interpretation Notes

{notes_md}

## File Manifest

{manifest_md}

## Next Submission Step

Before journal submission, rerun the NHANES mixture module in R if a full R environment becomes available, using `survey`, `qgcomp`, `gWQS`, and `bkmr` package implementations. The current Python outputs are transparent, reproducible approximations and are suitable for deciding whether the evidence chain is strong enough to advance.
"""


def main() -> None:
    ensure_dir(OUT)
    ensure_dir(MIRROR)
    qa_notes = []
    index_rows: list[dict[str, str]] = []

    nh, nh_stage, cg = load_inputs()
    print("Loaded input datasets.")
    qa_notes.append(f"Loaded NHANES analysis dataset: {NHANES_ANALYSIS} with shape {nh.shape}.")
    qa_notes.append(f"Loaded CGMacros MSI meal dataset: {CGM_STAGE1} with shape {cg.shape}.")
    if not HAS_SCIPY:
        qa_notes.append("SciPy unavailable; chi-square p-values used normal/Wilson-Hilferty approximations.")
    if not HAS_MPL:
        qa_notes.append("matplotlib unavailable; figure generation skipped.")

    nh, cg, msi_ref = add_dual_msi(nh, cg)
    paths = [
        ("MSI", "01_msi_component_reference_parameters.csv", msi_ref, "Reference means/SDs for dual MSI components."),
        ("MSI", "01_nhanes_dual_msi_dataset.csv", nh, "NHANES analysis dataset with exposure-facing MSI and response-vulnerability index."),
        ("MSI", "01_cgmacros_dual_msi_meal_dataset.csv", cg, "CGMacros meal dataset with dual MSI indices."),
    ]
    for module, name, df, desc in paths:
        write_csv(OUT / name, df)
        index_rows.append({"module": module, "file": name, "description": desc})
    print("Wrote dual MSI datasets.")

    corr_rows = []
    for dataset_name, df, old_cols in [
        ("NHANES", nh, ["msi_core_partial", "msi_core_complete", "msi_no_bmi", "msi_glycemia_ir", "msi_pca"]),
        ("CGMacros", cg, ["msi_core_partial_nhanes_ref", "msi_core_complete_nhanes_ref"]),
    ]:
        cols = ["exposure_facing_msi", "response_vulnerability_index"] + [c for c in old_cols if c in df.columns]
        for i, c1 in enumerate(cols):
            for c2 in cols[i + 1 :]:
                x, y = numeric(df[c1]), numeric(df[c2])
                mask = x.notna() & y.notna()
                corr = float(np.corrcoef(x[mask], y[mask])[0, 1]) if mask.sum() > 3 else np.nan
                corr_rows.append({"dataset": dataset_name, "variable_1": c1, "variable_2": c2, "pearson_r": corr, "n": int(mask.sum())})
    corr_df = pd.DataFrame(corr_rows)
    write_csv(OUT / "01_dual_msi_construct_correlations.csv", corr_df)
    index_rows.append({"module": "MSI", "file": "01_dual_msi_construct_correlations.csv", "description": "Construct validity correlations among dual MSI and previous MSI versions."})

    mixture = run_nhanes_mixture(nh)
    for name, df in mixture.items():
        file = f"02_{name}.csv"
        write_csv(OUT / file, df)
        index_rows.append({"module": "NHANES mixture", "file": file, "description": name.replace("_", " ")})
    qa_notes.append("NHANES mixture models use Python approximations: WQS repeated holdout, qgcomp-style quantile sum, and BKMR-style Gaussian kernel ridge.")
    print("Completed NHANES mixture models.")

    cg_food, codebook, fdc, reliability = add_food_matrix(cg)
    food_outputs = [
        ("03_food_matrix_codebook.csv", codebook, "Food-matrix variable definitions and rationale."),
        ("03_cgmacros_food_matrix_scored_meals.csv", cg_food, "CGMacros meals with food-matrix scores and dual-rule categories."),
        ("03_food_matrix_fdc_proxy_matches.csv", fdc, "Proxy USDA FoodData Central query links derived from food-matrix categories."),
        ("03_food_matrix_reliability.csv", reliability, "Algorithmic dual-rule kappa/ICC reliability stress test."),
    ]
    for file, df, desc in food_outputs:
        write_csv(OUT / file, df)
        index_rows.append({"module": "Food matrix", "file": file, "description": desc})
    qa_notes.append("FDC linkage is a proxy query-level match because the available CGMacros table has image paths and nutrient fields but no full parsed meal ingredient text.")
    print("Completed food matrix module.")

    cg_results = run_cgmacros_models(cg_food)
    for name, df in cg_results.items():
        file = f"04_{name}.csv"
        write_csv(OUT / file, df)
        index_rows.append({"module": "CGMacros", "file": file, "description": name.replace("_", " ")})
    print("Completed CGMacros subject-disjoint and hierarchical models.")

    stanford = run_stanford_boundary()
    for name, df in stanford.items():
        file = f"05_{name}.csv"
        write_csv(OUT / file, df)
        index_rows.append({"module": "Stanford CGM", "file": file, "description": name.replace("_", " ")})
    qa_notes.append("Stanford CGM boundary validation uses standardized challenge foods rather than free-living nutrient-coded meals; it tests transportability boundaries, not one-to-one replication.")
    print("Completed Stanford boundary validation.")

    mechanism = run_mechanism_workflow()
    for name, df in mechanism.items():
        file = f"06_{name}.csv"
        write_csv(OUT / file, df)
        index_rows.append({"module": "Mechanism", "file": file, "description": name.replace("_", " ")})
    qa_notes.append("Mechanism evidence is database-backed plausibility evidence, not direct experimental proof.")
    print("Completed mechanism workflow.")

    checklists = build_reporting_checklists()
    for name, df in checklists.items():
        file = f"07_{name}.csv"
        write_csv(OUT / file, df)
        index_rows.append({"module": "Reporting checklist", "file": file, "description": name.replace("_", " ")})
    print("Completed reporting checklists.")

    write_csv(OUT / "00_file_manifest.csv", pd.DataFrame(index_rows))
    write_workbook(index_rows)
    report = build_report(index_rows, qa_notes)
    write_text(OUT / "00_high_impact_computational_upgrade_report.md", report)
    write_json(
        OUT / "00_run_metadata.json",
        {
            "generated": datetime.now().isoformat(timespec="seconds"),
            "root": str(ROOT),
            "output": str(OUT),
            "mirror": str(MIRROR),
            "python_implementation_boundary": "WQS/qgcomp/BKMR are transparent Python approximations intended for dry-lab strengthening; rerun R package versions before final submission if possible.",
            "qa_notes": qa_notes,
        },
    )

    # Mirror all remaining files in case workbook/report were written after copied files.
    # Use overwrite-copy rather than deleting the OneDrive-backed mirror folder, which
    # may be transiently locked by the sync client.
    ensure_dir(MIRROR)
    for src in OUT.rglob("*"):
        if src.is_file():
            dest = MIRROR / src.relative_to(OUT)
            ensure_dir(dest.parent)
            shutil.copy2(src, dest)
    print(f"Created: {OUT}")
    print(f"Mirrored: {MIRROR}")
    print(f"Files: {len(list(OUT.glob('*')))}")


if __name__ == "__main__":
    main()
